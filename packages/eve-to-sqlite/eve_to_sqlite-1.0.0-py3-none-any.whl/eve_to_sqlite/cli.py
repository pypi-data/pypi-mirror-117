import click
import os
import zipfile
import sqlite_utils
from openpyxl import load_workbook

def write_records(records, db):
    # We write records into tables based on their types
    records_by_type = {}
    for record in records:
        table = record.pop("type").lower()
        records_by_type.setdefault(table, []).append(record)

    # Bulk inserts for each one
    for table, records_for_table in records_by_type.items():
        db[table].insert_all(
            records_for_table,
            alter=True,
            column_order=["date", "value", "accessory"],
            batch_size=50,
        )


@click.command()
@click.argument(
    "export_xlsx",
    type=click.Path(exists=True, file_okay=True, dir_okay=False, allow_dash=False),
    required=True,
    nargs=-1,
)
@click.argument(
    "db_path",
    type=click.Path(file_okay=True, dir_okay=False, allow_dash=False),
    required=True,
)
def cli(export_xlsx, db_path):
    "Convert Eve measurement exports to a SQLite database"

    db = sqlite_utils.Database(db_path)

    for file in export_xlsx:
        wb = load_workbook(filename = file)
        ws = wb.active
        type = ws.title

        accessory = None
        date_cell = None
        for row in ws.iter_rows(min_row=1, max_col=1, max_row=5):
            for cell in row:
                if cell.value.startswith("Accessory: "):
                    accessory = cell.value[len("Accessory: "):]

                if cell.value == "Date":
                    date_cell = cell
                    break
            if date_cell:
                break

        if date_cell is None:
            print(f"No date header cell found in {file}")
            continue

        date_iter = date_cell.offset(row=1)
        records = []
        while True:
            date_val = date_iter.value
            if date_val is None:
                break

            value_cell = date_iter.offset(column=1)
            value = value_cell.value
            date_val = date_val.strftime("%Y-%m-%dT%H:%M:%SZ")
            record = {
                "type": type,
                "date": date_val,
                "value": value,
                "accessory": accessory
            }
            records.append(record)

            date_iter = date_iter.offset(row=1)

            if len(records) >= 200:
                write_records(records, db)
                records = []

        if records:
            write_records(records, db)
