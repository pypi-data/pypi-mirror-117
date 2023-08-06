import os
from datetime import date
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from typing import Any, Dict, List


def get_workbook(filepath: str) -> Workbook:
    """get_workbook Open Workbook if file exist

    Args:
        filepath (str): File path

    Returns:
        Workbook: Workbook
    """
    if os.path.isfile(filepath):
        return load_workbook(filepath)
    return Workbook()


def get_data_ws(ws: Worksheet, mapping: Dict[str, str], row: int) -> Dict[str, Any]:
    """get_data_ws get data from worksheet

    Args:
        ws (Worksheet): Openpyxl worksheet
        mapping (Dict[str, str]): Key and collumn combination
        row (int): Row number

    Returns:
        Dict[str, Any]: Data?
    """
    result: Dict[str, Any] = dict()
    for key, col in mapping.items():
        result[key] = ws[f"{col}{row}"].value
    return result


def get_data_excel(
    filepath: str,
    sheet: str,
    rows: List[int],
    mapping: Dict[str, str],
) -> Dict[int, Dict[str, Any]]:
    """get_data_excel get data from excel file

    Args:
        filepath (str): File directory
        sheet (str): Worksheet name
        rows (List[int]): Number of row
        mapping (Dict[str, str]): Key and collumn combination

    Returns:
        Dict[int, Dict[str, Any]]: Data?
    """
    results: Dict[int, Dict[str, Any]] = dict()
    wb = load_workbook(filepath)
    ws = wb[sheet]
    for row in rows:
        results[row] = get_data_ws(ws, mapping, row)
    return results


def parse_range(value: str) -> List[int]:
    results: List[int] = list()
    if value.isdigit():
        results.append(int(value))
    elif "," in value:
        for val in value.split(","):
            # For 1,2
            results.extend(parse_range(val))
    elif "-" in value:
        start, end = value.split("-")
        if start.isdigit() and end.isdigit():
            if start == end:
                # For 1-1
                results.append(int(start))
            else:
                # For 1-2
                results.extend(range(int(start), int(end) + 1))
    return results


def snake_to_title(text: str) -> str:
    text = text.replace("_", " ")
    return text.title()


def parse_date(text: str, SEPARATOR: str = "") -> date:
    if SEPARATOR:
        pass
    elif "/" in text:
        SEPARATOR = "/"
    elif ":" in text:
        SEPARATOR = ":"
    else:
        SEPARATOR = " "
    tanggal, bulan, tahun = text.split(SEPARATOR)
    return date(int(tahun), int(bulan), int(tanggal))
