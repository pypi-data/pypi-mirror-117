# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.cell import Cell  # openpyxl == 3.0.7,需要去__init__文件中import cell

import os
from typing import Optional, Union, Generator, List, Tuple, Any

from excel_tools.base_reader import ExcelBaseObject


class XlsxReader(ExcelBaseObject):
    def __init__(self, name: str, path: Optional[str] = '', ignore_lines: Optional[int] = 0,
                 sheet: Union[int, str, None] = 0):
        """
        xlsx读取类，可以读取以下格式的表格：xlsx

        Args:
            name(str): 文件名
            path(str): 文件路径
            ignore_lines(int): 读取表格时,忽略读取的行数。
                例：row=3,ignore_line=2,那么最后读取的就是第row+ignore_lines=5行
            sheet(int,str): 需要读取的工作表(表名或表索引)
        """
        super(XlsxReader, self).__init__(name, path, ignore_lines)
        self._xl = openpyxl.load_workbook(os.path.join(self._path, self._name))
        if isinstance(sheet, str):
            self._sheet = self._xl[sheet]
        elif isinstance(sheet, int):
            self._sheet = self._xl[self._xl.sheetnames[sheet]]
        else:
            self._sheet = self._xl.active

    @property
    def sheet_name(self) -> str:
        """
        当前读取中的工作表名称

        Returns:
            工作表名称
        """
        return self._sheet.title

    def get_rows(self) -> Generator[Tuple[Cell], Any, None]:
        """ 获取所有列的单元格 """
        return self._sheet.rows

    def get_row(self, rowx: int) -> Tuple[Cell]:
        """
        获取对应行的所有单元格

        Args:
            rowx: 行数

        Returns:
            返回这一行所有单元格
        """
        return list(self.get_rows())[rowx - 1]

    def get_row_len(self, rowx: int) -> int:
        """
        获取对应行有效单元格的数量

        Args:
            rowx: 行数

        Returns:
            有效单元格数量
        """
        return len(self.get_row(rowx))

    def get_row_value(self, rowx: int, start_colx: Optional[int] = 0, end_colx: int = None) -> List[Any]:
        """
        获取表格内对应行中的所有单元格的数据

        Args:
            rowx: 行数
            start_colx: 左切片列数
            end_colx: 右切片列数

        Returns:
            返回这一行所有数据组成的列表
        """
        row = self.get_row(rowx)[start_colx:end_colx]
        return [cell.value for cell in row]

    def get_col(self, colx) -> Tuple[Cell]:
        """
        获取表格内对应列中的所有单元格

        Args:
            colx: 行数

        Returns:
            返回这一列所有数据组成的列表
        """
        return self._sheet[get_column_letter(colx)]

    def get_col_value(self, colx: int, start_rowx: Optional[int] = 0, end_rowx: Optional[int] = None) -> List[Any]:
        """
        获取表格内对应列中的所有单元格的数据

        Args:
            colx: 行数
            start_rowx: 左切片行数
            end_rowx: 右切片行数

        Returns:
            返回这一列所有数据组成的列表
        """
        col = self.get_col(colx)[start_rowx:end_rowx]
        return [cell.value for cell in col]

    def get_cell(self, rowx, cols) -> Cell:
        """
        获取单元格

        Args:
            rowx: 列数
            cols: 行数

        Returns:
            行列对应单元格
        """
        rowx = rowx + self._ignore_lines
        return self._sheet.cell(rowx, cols).value